import os
import pdfplumber
import openpyxl
import re
from datetime import datetime, timedelta
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Constante: jornada semanal fixa = 08:48:00  →  fração de dia Excel
# ---------------------------------------------------------------------------
JORNADA_SEMANAL_STR = "08:48:00"
JORNADA_SEMANAL_FRACTION = (8 * 3600 + 48 * 60) / 86400.0

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------
_thin_side = Side(style='thin')
THIN_BORDER = Border(left=_thin_side, right=_thin_side,
                     top=_thin_side, bottom=_thin_side)
WEEK_TOTAL_FILL = PatternFill(start_color="DDF2FF", end_color="DDF2FF", fill_type="solid")
MONTHLY_FILL   = PatternFill(start_color="EAF2F9", end_color="EAF2F9", fill_type="solid")
WEEKEND_FILL   = PatternFill(start_color="CBCBCB", end_color="CBCBCB", fill_type="solid")
HEADER_FILL    = PatternFill(start_color="929292", end_color="929292", fill_type="solid")

# ---------------------------------------------------------------------------
# Funções auxiliares
# ---------------------------------------------------------------------------

def is_within_date_range(date_str, start_date_str, end_date_str):
    fmt = "%d/%m/%Y"
    d  = datetime.strptime(date_str,       fmt)
    s  = datetime.strptime(start_date_str, fmt)
    e  = datetime.strptime(end_date_str,   fmt)
    return s <= d <= e


def time_str_to_timedelta(time_str):
    if time_str:
        h, m, s = map(int, time_str.split(':'))
        return timedelta(hours=h, minutes=m, seconds=s)
    return timedelta(0)


def timedelta_to_time_str(td):
    total_seconds = int(td.total_seconds())
    h, remainder = divmod(total_seconds, 3600)
    m, s = divmod(remainder, 60)
    return f"{h:02}:{m:02}:{s:02}"


def extract_data_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = "".join(page.extract_text() for page in pdf.pages)
    pattern = re.compile(
        r'(\d{2}/\d{2}/\d{4})\s*'
        r'(\d{2}:\d{2}:\d{2})?\s*'
        r'(\d{2}:\d{2}:\d{2})?\s*'
        r'(\d{2}:\d{2}:\d{2})?\s*'
        r'(\d{2}:\d{2}:\d{2})?\s*'
        r'(\d{2}:\d{2}:\d{2})?\s*'
        r'(\d{2}:\d{2}:\d{2})?\s*'
        r'(\d{2}:\d{2}:\d{2})?'
    )
    return pattern.findall(text)


day_translation = {
    "Monday":    "Segunda-feira",
    "Tuesday":   "Terça-feira",
    "Wednesday": "Quarta-feira",
    "Thursday":  "Quinta-feira",
    "Friday":    "Sexta-feira",
    "Saturday":  "Sábado",
    "Sunday":    "Domingo",
}


def _weekday_pt(date_obj):
    return day_translation.get(date_obj.strftime("%A"), date_obj.strftime("%A"))


empresa_nome = "MR ORGANIZAÇÃO CONTABIL LTDA"
empresa_cnpj = "CNPJ: 19.331.844/0001-43"

_PT_ABBR = ["jan", "fev", "mar", "abr", "mai", "jun",
            "jul", "ago", "set", "out", "nov", "dez"]

# ---------------------------------------------------------------------------
# Escrita de células helpers
# ---------------------------------------------------------------------------

def _cell(ws, row, col, value="", bold=False, fill=None,
          number_format=None, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.border = THIN_BORDER
    c.font = Font(bold=bold)
    c.alignment = Alignment(horizontal=align, vertical='center')
    if fill:
        c.fill = fill
    if number_format:
        c.number_format = number_format
    return c


# ---------------------------------------------------------------------------
# Escrita de linhas
# ---------------------------------------------------------------------------

def _write_day_row(ws, row, date_str, date_obj, match):
    """Escreve uma linha de dia (colunas A..J)."""
    is_weekend = date_obj.weekday() >= 5  # Sab=5, Dom=6
    fill = WEEKEND_FILL if is_weekend else None

    _cell(ws, row, 1,  date_str,                        fill=fill)
    _cell(ws, row, 2,  _weekday_pt(date_obj),           fill=fill)
    for j, val in enumerate(match[1:7]):
        _cell(ws, row, 3 + j, val if val else "",       fill=fill)

    if is_weekend:
        for col in range(7, 11):
            _cell(ws, row, col, 0, number_format="[h]:mm:ss", fill=fill)
    else:
        _cell(ws, row, 7, f"=F{row}-C{row}-(E{row}-D{row})",
              number_format="[h]:mm:ss", fill=fill)
        _cell(ws, row, 8, JORNADA_SEMANAL_FRACTION,
              number_format="[h]:mm:ss", fill=fill)
        _cell(ws, row, 9, f"=IF(G{row}>H{row},G{row}-H{row},0)",
              number_format="[h]:mm:ss", fill=fill)
        _cell(ws, row, 10, f"=IF(G{row}<H{row},H{row}-G{row},0)",
              number_format="[h]:mm:ss", fill=fill)


def _write_weekly_total(ws, row, start_row, end_row):
    """Escreve a linha TOTAL semanal (bold, azul claro)."""
    for col in range(1, 11):
        _cell(ws, row, col, bold=True, fill=WEEK_TOTAL_FILL)

    ws.cell(row=row, column=1).value = "TOTAL"

    for col, letter in [(7, "G"), (8, "H"), (9, "I"), (10, "J")]:
        c = ws.cell(row=row, column=col)
        c.value = f"=SUM({letter}{start_row}:{letter}{end_row})"
        c.number_format = "[h]:mm:ss"
        c.font = Font(bold=True)


def _write_monthly_totals(ws, row, weekly_total_rows):
    """Escreve as 3 linhas de totais mensais."""
    if not weekly_total_rows:
        return

    g_refs = "+".join(f"G{r}" for r in weekly_total_rows)
    i_refs = "+".join(f"I{r}" for r in weekly_total_rows)
    j_refs = "+".join(f"J{r}" for r in weekly_total_rows)

    specs = [
        ("TOTAL MENSAL",    10,  g_refs),
        ("TOTAL POSITIVAS", 10,  i_refs),
        ("TOTAL NEGATIVAS", 10, j_refs),
    ]
    for label, value_col, refs in specs:
        for col in range(1, 11):
            _cell(ws, row, col, bold=True, fill=MONTHLY_FILL)
        ws.cell(row=row, column=1).value = label
        c = ws.cell(row=row, column=value_col)
        c.value = f"={refs}"
        c.number_format = "[h]:mm:ss"
        c.font = Font(bold=True)
        row += 1

    return row  # última linha usada


# ---------------------------------------------------------------------------
# Cabeçalho e info fixa
# ---------------------------------------------------------------------------

def _write_header(ws, person_name):
    """Linha 1 = nome da pessoa (merge A1:J1)."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    c = ws.cell(row=1, column=1, value=f"Relatório de Ponto {person_name}")
    c.font = Font(bold=True, size=14)
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = THIN_BORDER


def _write_column_headers(ws):
    """Linha 2 = cabeçalhos das colunas."""
    headers = [
        'Data', 'Dia da Semana',
        'Entrada (Manhã)', 'Saída (Manhã)',
        'Entrada (Tarde)', 'Saída (Tarde)',
        'Horas de Trabalho', 'H. Semanal',
        'Horas Extras', 'Horas Negativas',
    ]
    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=title)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = THIN_BORDER


def _write_signature_block(ws, person_name, cpf_val, start_row=42):
    """Bloco de assinatura e empresa."""
    r = start_row
    for col in range(7, 11):
        ws.cell(row=r,   column=col).border = THIN_BORDER
        ws.cell(row=r+1, column=col).border = THIN_BORDER
        ws.cell(row=r+2, column=col).border = THIN_BORDER
        ws.cell(row=r+4, column=col).border = THIN_BORDER
        ws.cell(row=r+5, column=col).border = THIN_BORDER

    font_sig = Font(name='Times New Roman', size=10)

    ws.merge_cells(start_row=r,   start_column=7, end_row=r,   end_column=10)
    ws.merge_cells(start_row=r+1, start_column=7, end_row=r+1, end_column=10)
    ws.merge_cells(start_row=r+2, start_column=7, end_row=r+2, end_column=10)
    ws.merge_cells(start_row=r+4, start_column=7, end_row=r+4, end_column=10)
    ws.merge_cells(start_row=r+5, start_column=7, end_row=r+5, end_column=10)

    ws.cell(row=r,   column=7, value="______________________________________").alignment = Alignment(horizontal='center')
    c_name = ws.cell(row=r+1, column=7, value=person_name.upper())
    c_name.font = font_sig; c_name.alignment = Alignment(horizontal='left')
    c_cpf = ws.cell(row=r+2, column=7, value=f"CPF: {cpf_val}")
    c_cpf.font = font_sig; c_cpf.alignment = Alignment(horizontal='left')
    c_emp = ws.cell(row=r+4, column=7, value=f" {empresa_nome}")
    c_emp.font = font_sig; c_emp.alignment = Alignment(horizontal='left')
    c_cnpj = ws.cell(row=r+5, column=7, value=f" {empresa_cnpj}")
    c_cnpj.font = font_sig; c_cnpj.alignment = Alignment(horizontal='left')


def _write_hour_statement(ws, start_row=48):
    """Extrato anual: 12 meses do ano corrente."""
    thin = THIN_BORDER
    ano_corrente = datetime.now().year
    ano_abbr = ano_corrente % 100

    headers_ext = ['Extrato hora', 'Horas Positivas - B.H',
                   'Horas Negativas', 'PG em Folha', 'Saldo']
    for col, title in enumerate(headers_ext, start=1):
        c = ws.cell(row=start_row, column=col, value=title)
        c.font = Font(bold=True); c.border = thin
        c.alignment = Alignment(horizontal='left', vertical='center')

    for i, abbr in enumerate(_PT_ABBR):
        r = start_row + 1 + i
        label = f"{abbr}-{ano_abbr:02d}"
        for col in range(1, 6):
            c = ws.cell(row=r, column=col, value=label if col == 1 else "")
            c.font = Font(bold=True); c.border = thin
            c.alignment = Alignment(horizontal='left', vertical='center')

    saldo_row = start_row + 13
    for col in range(1, 6):
        c = ws.cell(row=saldo_row, column=col, value="Saldo final" if col == 1 else "")
        c.font = Font(bold=True); c.border = thin
        c.alignment = Alignment(horizontal='left', vertical='center')


# ---------------------------------------------------------------------------
# Função principal de processamento
# ---------------------------------------------------------------------------

def process_pdfs(pdf_folder, output_folder, start_date, end_date, cpf_dict,
                 progress_callback=None):
    """
    Processa todos os PDFs em pdf_folder e gera planilhas Excel em output_folder.

    Layout de colunas por linha de dia:
      A  Data
      B  Dia da Semana
      C  Entrada (Manhã)
      D  Saída (Manhã)
      E  Entrada (Tarde)
      F  Saída (Tarde)
      G  Horas de Trabalho   = F-C-(E-D)
      H  H. Semanal          = 08:48:00 (fixo)
      I  Horas Extras        = IF(G>H, G-H, 0)
      J  Horas Negativas     = IF(G<H, H-G, 0)

    Após cada grupo de segunda a domingo: linha TOTAL semanal (bold, azul).
    Ao final: TOTAL MENSAL / TOTAL POSITIVAS / TOTAL NEGATIVAS.
    """
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    pdf_files = [f for f in os.listdir(pdf_folder) if f.lower().endswith('.pdf')]
    total_files = len(pdf_files)
    generated = []

    for idx, pdf_file in enumerate(pdf_files, start=1):
        pdf_path = os.path.join(pdf_folder, pdf_file)
        person_name = os.path.splitext(pdf_file)[0].replace('Relatório de Ponto - ', '')

        data = extract_data_from_pdf(pdf_path)
        filtered_data = [m for m in data if is_within_date_range(m[0], start_date, end_date)]

        wb = openpyxl.Workbook()
        ws = wb.active

        # ── Linha 1: nome | Linha 2: cabeçalhos ──────────────────────────
        _write_header(ws, person_name)
        _write_column_headers(ws)

        # ── Corpo: dias agrupados por semana ISO (seg→dom) ────────────────
        current_row    = 3          # primeira linha de dado
        current_week   = None       # número ISO da semana atual
        week_start_row = 3          # primeira linha de dado da semana corrente
        weekly_total_rows = []      # linhas dos totais semanais (para total mensal)

        for match in filtered_data:
            date_str = match[0]
            date_obj = datetime.strptime(date_str, "%d/%m/%Y")
            week_num = date_obj.isocalendar()[1]   # número ISO de semana

            # Virou a semana → grava total da semana anterior
            if current_week is not None and week_num != current_week:
                _write_weekly_total(ws, current_row, week_start_row, current_row - 1)
                weekly_total_rows.append(current_row)
                current_row    += 1
                week_start_row  = current_row

            _write_day_row(ws, current_row, date_str, date_obj, match)
            current_week  = week_num
            current_row  += 1

        # Fecha a última semana
        if current_week is not None and week_start_row < current_row:
            _write_weekly_total(ws, current_row, week_start_row, current_row - 1)
            weekly_total_rows.append(current_row)
            current_row += 1

        # ── Linha em branco + totais mensais ─────────────────────────────
        current_row += 1  # linha em branco de separação
        _write_monthly_totals(ws, current_row, weekly_total_rows)

        # ── Larguras de coluna ────────────────────────────────────────────
        col_widths = [15, 20, 18, 18, 18, 18, 20, 18, 18, 18]
        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        # Formata colunas de tempo (C a J = cols 3..10)
        for col in range(3, 11):
            col_letter = openpyxl.utils.get_column_letter(col)
            for cell in ws[col_letter]:
                if "[h]" in (cell.number_format or ""):
                    pass   # já formatado
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # ── Bloco de assinatura e extrato horário ─────────────────────────
        cpf_val = cpf_dict.get(person_name.strip().lower(), "000.000.000-00")
        cpf_val = cpf_val.replace("CPF: ", "")
        _write_signature_block(ws, person_name, cpf_val, start_row=42)
        _write_hour_statement(ws, start_row=48)

        # ── Salva ─────────────────────────────────────────────────────────
        excel_filename = f'Relatório de Ponto {person_name}.xlsx'
        output_path = os.path.join(output_folder, excel_filename)
        wb.save(output_path)
        generated.append(output_path)

        if progress_callback:
            progress_callback(int(idx / total_files * 100), f"Processado: {pdf_file}")

    return generated
