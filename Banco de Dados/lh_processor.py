# lh_processor.py
# -*- coding: utf-8 -*-

import os
import re
import calendar
import unicodedata
from datetime import date
from typing import Callable, Dict, Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ===== Saída =====
OUT_FILENAME = "relatorios_modelo_todos_gerado.xlsx"

# ===== Estilo =====
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
WEEK_TOTAL_FILL = PatternFill("solid", fgColor="DDF2FF")
WEEKEND_FILL = PatternFill("solid", fgColor="E9F6FD")
GRAY_FILL = PatternFill("solid", fgColor="EAF2F9")

# ===== Meses PT-BR =====
_PT_MONTHS = {
    1: "JANEIRO", 2: "FEVEREIRO", 3: "MARÇO", 4: "ABRIL",
    5: "MAIO", 6: "JUNHO", 7: "JULHO", 8: "AGOSTO",
    9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO", 12: "DEZEMBRO",
}
_PT_MONTHS_INV = {
    "JANEIRO": 1, "FEVEREIRO": 2, "MARÇO": 3, "MARCO": 3, "ABRIL": 4, "MAIO": 5, "JUNHO": 6,
    "JULHO": 7, "AGOSTO": 8, "SETEMBRO": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12
}
_PT_ABBR = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]

# ===== Soma compatível (localidade) =====
# Defina LH_SUM_MODE=['sum','soma','plus'] para escolher: =SUM, =SOMA, ou somatório com '+'
_SUM_MODE = os.getenv("LH_SUM_MODE", "plus").strip().lower()

# ===== Identificação fixa =====
EMP_CPF_FIXED = "000.000.000-00"
COMPANY_NAME = "COLEGIO ELLEVE LTDA"
COMPANY_CNPJ = "47.136.937/0001-52"

# ===== Configuração fixa por colaborador (NOME -> CPF + jornada semanal) =====
# week_times = [Seg, Ter, Qua, Qui, Sex, Sáb, Dom] em "HH:MM" ou "HH:MM:SS"
# ATENÇÃO: ajuste os nomes, CPFs e horários conforme sua realidade.

# week_times: [segunda, terça, quarta, quinta, sexta, sábado, domingo]

PREDEFINED_PEOPLE: Dict[str, Dict[str, Any]] = {
    "ALINE": {
        "cpf": "384.220.068-47",
        # segunda e terça: 8h | quarta e quinta: 5h
        "week_times": ["08:00:00", "08:00:00", "05:00:00", "05:00:00", "00:00:00", "00:00:00", "00:00:00"],
    },
    "AMANDA": {
        "cpf": "476.675.518-98",
        "week_times": ["05:00:00", "05:00:00", "05:00:00", "05:00:00", "05:00:00", "00:00:00", "00:00:00"],
    },
    "ARIANA": {
        "cpf": "340.305.828.00",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "ARIANE": {
        "cpf": "412.790.448-88",
        "week_times": ["05:00:00", "05:00:00", "05:00:00", "05:00:00", "05:00:00", "00:00:00", "00:00:00"],
    },
    "BARBARA": {
        "cpf": "446.285.768-23",
        "week_times": ["04:00:00", "04:00:00", "04:00:00", "04:00:00", "04:00:00", "00:00:00", "00:00:00"],
    },
    "BEATRIZ": {
        "cpf": "473.112.208-23",
        "week_times": ["05:00:00", "05:00:00", "05:00:00", "05:00:00", "05:00:00", "00:00:00", "00:00:00"],
    },
    "BRUNA": {
        "cpf": "389.180.268-46",
        "week_times": ["05:00:00", "05:00:00", "05:00:00", "05:00:00", "05:00:00", "00:00:00", "00:00:00"],
    },
    "CELESTINA": {
        "cpf": "284.364.338-40",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "CONCEICAO": {
        "cpf": "621.646.601-10",
        "week_times": ["05:00:00", "05:00:00", "05:00:00", "05:00:00", "05:00:00", "00:00:00", "00:00:00"],
    },
    "DAIANA": {
        "cpf": "337.572.508-69",
        "week_times": ["06:00:00", "06:00:00", "06:00:00", "06:00:00", "06:00:00", "00:00:00", "00:00:00"],
    },
    "DILMA": {
        "cpf": "296.882.348-50",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "EDRA": {
        "cpf": "169.273.728-73",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "EMILIA": {
        "cpf": "435.688.288-12",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "EMILLY": {
        "cpf": "554.291.848-45",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "ERICA": {
        "cpf": "332.987.088-52",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "GERLANDIA": {
        "cpf": "021.440.483-80",
        "week_times": ["05:00:00", "05:00:00", "05:00:00", "05:00:00", "05:00:00", "00:00:00", "00:00:00"],
    },
    "GIOVANA": {
        "cpf": "537.771.808-90",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "GUILHERME": {
        "cpf": "485.038.738-18",
        # segunda e quarta: 8:48h | terça e quinta: 5h
        "week_times": ["08:48:00", "05:00:00", "08:48:00", "05:00:00", "00:00:00", "00:00:00", "00:00:00"],
    },
    "HILDA": {
        "cpf": "113.635.908-71",
        "week_times": ["05:00:00", "05:00:00", "05:00:00", "05:00:00", "05:00:00", "00:00:00", "00:00:00"],
    },
    "ISADORA": {
        "cpf": "500.263.528-85",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "ISAQUE": {
        "cpf": "334.020.512-49",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "JOSELIA": {
        "cpf": "026.085.093-47",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "JULIANE": {
        "cpf": "504.586.078-00",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "KAIENE": {
        "cpf": "433.526.048-27",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "KAMILLA": {
        "cpf": "544.958.058-43",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "LETICIA": {
        "cpf": "504.174.088-70",
        "week_times": ["05:00:00", "05:00:00", "05:00:00", "05:00:00", "05:00:00", "00:00:00", "00:00:00"],
    },
    "MARIA": {
        "cpf": "145.141.228-23",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "MARIANE": {
        "cpf": "311.754.918-88",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "MONICA": {
        "cpf": "437.780.908-36",
        "week_times": ["05:00:00", "05:00:00", "05:00:00", "05:00:00", "05:00:00", "00:00:00", "00:00:00"],
    },
    "PAOLA": {
        "cpf": "108.247.037-66",
        "week_times": ["05:00:00", "05:00:00", "05:00:00", "05:00:00", "05:00:00", "00:00:00", "00:00:00"],
    },
    "PATRICIA": {
        "cpf": "103.210.298-52",
        "week_times": ["05:00:00", "05:00:00", "05:00:00", "05:00:00", "05:00:00", "00:00:00", "00:00:00"],
    },
    "PAULA": {
        "cpf": "362.855.388-16",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "PRISCILA": {
        "cpf": "321.504.628-85",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "RODRIGO": {
        "cpf": "350.986.768-86",
        "week_times": ["05:00:00", "05:00:00", "05:00:00", "05:00:00", "05:00:00", "00:00:00", "00:00:00"],
    },
    "STEPHANY": {
        "cpf": "479.320.558-57",
        # Em licença maternidade
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
    "THAMIRIS": {
        "cpf": "423.041.298-38",
        # terça: 8h | quinta: 5h
        "week_times": ["00:00:00", "08:00:00", "00:00:00", "05:00:00", "00:00:00", "00:00:00", "00:00:00"],
    },
    "VIVIANE": {
        "cpf": "326.858.838-78",
        "week_times": ["08:48:00", "08:48:00", "08:48:00", "08:48:00", "08:48:00", "00:00:00", "00:00:00"],
    },
}


# ------------------ utilitários ------------------


def _sum_range_formula(col_letter: str, r1: int, r2: int) -> str:
    """Gera fórmula de soma vertical conforme LH_SUM_MODE."""
    if r2 < r1:
        return "0"
    if _SUM_MODE == "sum":
        return f"=SUM({col_letter}{r1}:{col_letter}{r2})"
    if _SUM_MODE == "soma":
        return f"=SOMA({col_letter}{r1}:{col_letter}{r2})"
    terms = [f"{col_letter}{r}" for r in range(r1, r2 + 1)]
    return "=" + "+".join(terms) if terms else "0"


def _norm(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.strip().upper()


def _person_cfg(nome: str) -> Optional[Dict[str, Any]]:
    """Retorna o dicionário de config fixa para o colaborador (se existir)."""
    key = _norm(nome)
    return PREDEFINED_PEOPLE.get(key)


def _parse_times(cell) -> tuple:
    """Extrai até 4 horários hh:mm; retorna (E-Manhã, S-Manhã, E-Tarde, S-Tarde)."""
    if pd.isna(cell):
        return "", "", "", ""
    s = str(cell).replace("\\n", "\n")
    times = re.findall(r"(\d{1,2}:\d{2})", s)
    times = [f"{int(t.split(':')[0]):02d}:{t.split(':')[1]}" for t in times]
    if len(times) >= 4:
        return times[0], times[1], times[2], times[3]
    if len(times) == 3:
        return times[0], times[1], times[2], ""
    if len(times) == 2:
        return times[0], "", "", times[1]
    if len(times) == 1:
        return times[0], "", "", ""
    return "", "", "", ""


def _weekday_pt(dt: date) -> str:
    return ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira",
            "Sexta-feira", "Sábado", "Domingo"][dt.weekday()]


def _month_year_from_c2(df) -> tuple | None:
    """Detecta (mês, ano) exclusivamente da célula C2 (diversos formatos)."""
    try:
        c2 = df.iat[1, 2]  # C2
    except Exception:
        c2 = ""
    raw = str(c2)
    txt = _norm(raw)

    # mm/yyyy
    m = re.search(r"\b(\d{1,2})\s*[/\-.]\s*(\d{4})\b", txt)
    if m:
        mm = int(m.group(1))
        yyyy = int(m.group(2))
        if 1 <= mm <= 12:
            return mm, yyyy

    # OUTUBRO 2025 etc
    m = re.search(r"\b([A-ZÇÃÉÓ]+)\s*[/\-\s]*\s*(\d{4})\b", txt)
    if m:
        mon = (m.group(1)
               .replace("Ç", "C").replace("Ã", "A")
               .replace("É", "E").replace("Ó", "O"))
        yyyy = int(m.group(2))
        if mon in _PT_MONTHS_INV:
            return _PT_MONTHS_INV[mon], yyyy

    # range 01.10.2025 a 31.10.2025
    m = re.search(r"(\d{2})\.(\d{2})\.(\d{4}).*?(\d{2})\.(\d{2})\.(\d{4})", raw)
    if m:
        return int(m.group(2)), int(m.group(3))

    # dd.mm.yyyy
    m = re.search(r"\b\d{2}\.(\d{2})\.(\d{4})\b", raw)
    if m:
        return int(m.group(1)), int(m.group(2))

    return None


def _hms_to_fraction(h: int, m: int, s: int = 0) -> float:
    """Converte horas/min/seg em fração de dia (padrão Excel)."""
    total_seg = h * 3600 + m * 60 + s
    return total_seg / 86400.0


def _time_str_to_fraction(s: str) -> float:
    """
    Converte 'HH:MM' ou 'HH:MM:SS' em fração de dia.
    Se vazio ou inválido, retorna 0.
    """
    s = (s or "").strip()
    if not s:
        return 0.0
    try:
        parts = s.split(":")
        if len(parts) == 2:
            h, m = int(parts[0]), int(parts[1])
            return _hms_to_fraction(h, m, 0)
        if len(parts) == 3:
            h, m, sec = int(parts[0]), int(parts[1]), int(parts[2])
            return _hms_to_fraction(h, m, sec)
    except Exception:
        return 0.0
    return 0.0


# ------------------ preview (GUI) ------------------


def quick_preview(file_path: str) -> Dict:
    xls = pd.ExcelFile(file_path)
    df = pd.read_excel(file_path, sheet_name="Registro de atendimento", header=None)
    n_rows, n_cols = df.shape

    names = []
    for r in range(2, n_rows, 4):
        if 11 >= n_cols:
            break
        try:
            val = df.iat[r, 11]
        except Exception:
            continue
        if pd.isna(val):
            continue
        names.append(str(val).strip())

    det = _month_year_from_c2(df)
    month_year = f"{det[0]:02d}/{det[1]}" if det else None

    return {"names": names, "month_year": month_year, "sheets": list(xls.sheet_names)}


# ------------------ gerador principal ------------------


def process_file(
    in_file: str,
    out_dir: str,
    progress_callback: Callable[[int, str], None] = None,
    target_hours_provider: Optional[Callable[[str, date], Optional[float]]] = None,
) -> str:
    """
    Gera o workbook final a partir da aba 'Registro de atendimento'.

    - target_hours_provider(nome: str, data: date) -> horas (float) ou None
      * Se informado: retorna horas-alvo do dia (ex.: 7.0, 8.5). Convertemos para dias (Excel = horas/24).
      * Se None ou não definir nada para o dia, usamos:
          1) Se existir em PREDEFINED_PEOPLE: week_times[weekday] em HH:MM[:SS]
          2) Caso contrário: fallback 8h em dias úteis e 0h em finais de semana.
    """
    if progress_callback is None:
        progress_callback = lambda p, m="": None

    # Lê apenas a aba solicitada
    df = pd.read_excel(in_file, sheet_name="Registro de atendimento", header=None)
    n_rows, n_cols = df.shape

    det = _month_year_from_c2(df)
    if not det:
        raise ValueError("Não foi possível identificar mês/ano a partir da célula C2 da planilha base.")
    mes, ano = det
    num_days = calendar.monthrange(ano, mes)[1]

    wb = Workbook()
    tmp = wb.active
    tmp.title = "tmp"

    # Mapeia nomes válidos (coluna L nas linhas 2,6,10,...) e ordena alfabeticamente
    name_rows = list(range(2, n_rows, 4))
    valid_names = []
    for r in name_rows:
        if r >= n_rows or 11 >= n_cols:
            continue
        nome = df.iat[r, 11]
        if pd.isna(nome) or str(nome).strip() == "":
            continue
        valid_names.append((str(nome).strip(), r))
    # Ordenação alfabética (case-insensitive)
    valid_names.sort(key=lambda x: x[0].lower())
    total_names = len(valid_names)

    processed = 0
    for nome, r_name in valid_names:
        r_days = r_name + 1
        r_hours = r_name + 3
        if r_hours >= n_rows:
            continue

        # Até 31 colunas no modelo; usamos até o número real de dias
        days = df.iloc[r_days, :31].reset_index(drop=True)   # mantido por compatibilidade
        hours = df.iloc[r_hours, :31].reset_index(drop=True)

        ws = wb.create_sheet(title=nome[:31])

        # ---------- Cabeçalho ----------
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=11)
        titulo_mes = _PT_MONTHS.get(mes, date(ano, mes, 1).strftime('%B').upper())
        ws.cell(row=1, column=2, value=f"{nome} - {titulo_mes} / {ano}").font = Font(bold=True, size=14)
        ws.cell(row=1, column=2).alignment = center

        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=11)
        ws.cell(row=2, column=2, value="HORÁRIO: ________________________________").alignment = left

        headers = [
            "Data", "Dia da Semana", "Entrada - Manhã", "Saída (Manhã)",
            "Entrada (Tarde)", "Saída (Tarde)", "Horas Trabalhadas",
            "H. Semanal", "H. Extras", "H. Negativas"
        ]
        for j, h in enumerate(headers, start=2):
            c = ws.cell(row=3, column=j, value=h)
            c.font = Font(bold=True)
            c.fill = HEADER_FILL
            c.alignment = center
            c.border = border

        # Helper: jornada-alvo por dia (em fração do dia p/ Excel)
        def _target_for(person_name: str, dt_obj: date) -> float:
            # 1) Provider externo (JSON / diálogo), se houver
            if target_hours_provider:
                try:
                    val = target_hours_provider(person_name or "", dt_obj)
                    if isinstance(val, (int, float)) and val >= 0:
                        return float(val) / 24.0
                except Exception:
                    pass

            # 2) Configuração fixa no código (PREDEFINED_PEOPLE)
            cfg = _person_cfg(person_name)
            if cfg:
                week_times = cfg.get("week_times")
                if week_times and 0 <= dt_obj.weekday() < len(week_times):
                    return _time_str_to_fraction(week_times[dt_obj.weekday()])

                # opcional: suporte a 'week_hours' em horas (caso queira misturar)
                week_hours = cfg.get("week_hours")
                if week_hours and 0 <= dt_obj.weekday() < len(week_hours):
                    try:
                        return float(week_hours[dt_obj.weekday()]) / 24.0
                    except Exception:
                        pass

            # 3) Fallback antigo: 8h em dias úteis; 0h em finais de semana
            return (8.8 / 24.0) if dt_obj.weekday() < 5 else 0.0

        def _cpf_for(person_name: str) -> str:
            cfg = _person_cfg(person_name)
            if cfg and cfg.get("cpf"):
                return str(cfg["cpf"])
            return EMP_CPF_FIXED

        # ---------- Corpo diário + Totais semanais ----------
        row = 4
        week_start = None
        weekly_totals = []

        for d in range(1, num_days + 1):
            dt = date(ano, mes, d)

            # Fecha semana anterior quando vira segunda
            if dt.weekday() == 0 and week_start is not None:
                # Linha TOTAL semanal
                for cidx in range(2, 12):
                    ws.cell(row=row, column=cidx).fill = WEEK_TOTAL_FILL
                    ws.cell(row=row, column=cidx).border = border
                    ws.cell(row=row, column=cidx).alignment = center
                    ws.cell(row=row, column=cidx).font = Font(bold=True)

                ws.cell(row=row, column=2, value="TOTAL")
                ws.cell(row=row, column=8, value=_sum_range_formula("H", week_start, row - 1)).number_format = "[h]:mm:ss"
                ws.cell(row=row, column=9, value=None)
                ws.cell(row=row, column=10, value=_sum_range_formula("J", week_start, row - 1)).number_format = "[h]:mm:ss"
                ws.cell(row=row, column=11, value=_sum_range_formula("K", week_start, row - 1)).number_format = "[h]:mm:ss"

                weekly_totals.append(row)
                row += 1
                week_start = None

            # Linha do dia
            ws.cell(row=row, column=2, value=dt.strftime("%d/%b").lower()).alignment = center
            ws.cell(row=row, column=2).border = border
            ws.cell(row=row, column=3, value=_weekday_pt(dt)).alignment = center
            ws.cell(row=row, column=3).border = border

            hour_cell = hours.iloc[d - 1] if (d - 1) < len(hours) else ""
            em, sm, et, st = _parse_times(hour_cell)
            ws.cell(row=row, column=4, value=em).alignment = center
            ws.cell(row=row, column=4).border = border
            ws.cell(row=row, column=5, value=sm).alignment = center
            ws.cell(row=row, column=5).border = border
            ws.cell(row=row, column=6, value=et).alignment = center
            ws.cell(row=row, column=6).border = border
            ws.cell(row=row, column=7, value=st).alignment = center
            ws.cell(row=row, column=7).border = border

            # H = Horas trabalhadas
            ws.cell(row=row, column=8, value=f"=(G{row}-D{row})-(F{row}-E{row})").number_format = "[h]:mm:ss"
            ws.cell(row=row, column=8).alignment = center
            ws.cell(row=row, column=8).border = border

            # I = Jornada-alvo (provider/definição fixa -> fração do dia)
            i_value = _target_for(nome, dt)
            ws.cell(row=row, column=9, value=i_value).number_format = "[h]:mm:ss"
            ws.cell(row=row, column=9).alignment = center
            ws.cell(row=row, column=9).border = border

            # J = Horas Extras (positivas)
            ws.cell(row=row, column=10, value=f"=IF(H{row}>I{row}, H{row}-I{row}, 0)").number_format = "[h]:mm:ss"
            ws.cell(row=row, column=10).alignment = center
            ws.cell(row=row, column=10).border = border

            # K = Horas Negativas
            ws.cell(row=row, column=11, value=f"=IF(H{row}<I{row}, I{row}-H{row}, 0)").number_format = "[h]:mm:ss"
            ws.cell(row=row, column=11).alignment = center
            ws.cell(row=row, column=11).border = border

            # Finais de semana com preenchimento
            if dt.weekday() >= 5:
                for cidx in range(2, 12):
                    ws.cell(row=row, column=cidx).fill = WEEKEND_FILL

            if week_start is None:
                week_start = row
            row += 1

        # Fecha última semana (se aberta)
        if week_start is not None and row > week_start:
            for cidx in range(2, 12):
                ws.cell(row=row, column=cidx).fill = GRAY_FILL
                ws.cell(row=row, column=cidx).border = border
                ws.cell(row=row, column=cidx).alignment = center
                ws.cell(row=row, column=cidx).font = Font(bold=True)

            ws.cell(row=row, column=2, value="TOTAL")
            ws.cell(row=row, column=8, value=_sum_range_formula("H", week_start, row - 1)).number_format = "[h]:mm:ss"
            ws.cell(row=row, column=9, value=None)
            ws.cell(row=row, column=10, value=_sum_range_formula("J", week_start, row - 1)).number_format = "[h]:mm:ss"
            ws.cell(row=row, column=11, value=_sum_range_formula("K", week_start, row - 1)).number_format = "[h]:mm:ss"

            weekly_totals.append(row)
            row += 1

        # ---------- Totais finais (linhas 40..42) ----------
        labels = {40: "TOTAL MENSAL", 41: "TOTAL POSITIVAS", 42: "TOTAL NEGATIVAS"}
        for rlab, txt in labels.items():
            for cidx in range(2, 12):
                cell = ws.cell(row=rlab, column=cidx)
                cell.fill = GRAY_FILL
                cell.border = border
                cell.alignment = center
            ws.cell(row=rlab, column=2, value=txt).font = Font(bold=True)
            ws.cell(row=rlab, column=11).font = Font(bold=True)
            ws.cell(row=rlab, column=11).number_format = "[h]:mm:ss"

        if weekly_totals:
            refs_H = [f"H{r}" for r in weekly_totals]
            refs_J = [f"J{r}" for r in weekly_totals]
            refs_K = [f"K{r}" for r in weekly_totals]
            if refs_H:
                ws.cell(row=40, column=11, value="=" + "+".join(refs_H)).number_format = "[h]:mm:ss"
            if refs_J:
                ws.cell(row=41, column=11, value="=" + "+".join(refs_J)).number_format = "[h]:mm:ss"
            if refs_K:
                ws.cell(row=42, column=11, value="=" + "+".join(refs_K)).number_format = "[h]:mm:ss"

        # >>> Bordas completas no bloco principal (B1:K42)
        for rr in range(1, 43):
            for cc in range(2, 12):
                ws.cell(row=rr, column=cc).border = border

        # ---------- Extrato anual (B..F) ----------
        base = 44
        extrato_headers = ["Extrato hora", "Horas Positivas - B.H", "Horas Negativas", "PG em Folha", "Saldo"]
        for col_off, title in enumerate(extrato_headers, start=2):
            c = ws.cell(row=base, column=col_off, value=title)
            c.font = Font(bold=True)
            c.fill = HEADER_FILL
            c.alignment = center
            c.border = border

        for i in range(12):
            r = base + 1 + i
            mcell = ws.cell(row=r, column=2, value=f"{_PT_ABBR[i]}-{ano % 100:02d}")
            mcell.alignment = center
            mcell.border = border
            for col in (3, 4, 5):
                vcell = ws.cell(row=r, column=col, value=0)
                vcell.number_format = "0.00"
                vcell.alignment = center
                vcell.border = border
            scell = ws.cell(row=r, column=6, value=f"=C{r}-D{r}-E{r}")
            scell.number_format = "0.00"
            scell.alignment = center
            scell.border = border

        # SALDO FINAL com preenchimento
        saldo_final_row = base + 13
        for col in range(2, 7):
            cell = ws.cell(row=saldo_final_row, column=col)
            cell.fill = GRAY_FILL
            cell.border = border
            cell.alignment = center
        ws.cell(row=saldo_final_row, column=2, value="SALDO FINAL").font = Font(bold=True)
        saldo_terms = [f"F{base + 1 + i}" for i in range(12)]
        ws.cell(row=saldo_final_row, column=6, value=("=" + "+".join(saldo_terms))).number_format = "0.00"

        # >>> Bordas completas no extrato (B44:F57)
        for rr in range(44, 58):
            for cc in range(2, 7):
                ws.cell(row=rr, column=cc).border = border

        # ---------- Bloco de assinatura (fixo a partir da linha 48 em H..K) ----------
        info_c1, info_c2 = 8, 11  # H..K

        def _merge_write(r: int, text: str):
            ws.merge_cells(start_row=r, start_column=info_c1, end_row=r, end_column=info_c2)
            c = ws.cell(row=r, column=info_c1, value=text)
            c.alignment = center
            c.font = Font(bold=True)

        SIGN_START_ROW = 48
        r_info = SIGN_START_ROW
        _merge_write(r_info, "______________________________________")
        r_info += 1
        _merge_write(r_info, nome.upper())
        r_info += 1
        cpf_val = _cpf_for(nome)
        _merge_write(r_info, f"CPF: {cpf_val}")
        r_info += 2
        _merge_write(r_info, COMPANY_NAME)
        r_info += 1
        _merge_write(r_info, f"CNPJ: {COMPANY_CNPJ}")

        # ---------- Larguras ----------
        widths = {
            2: 22,  # B (mais espaçada)
            3: 20,  # C (mais espaçada)
            4: 18,  # D (mais espaçada)
            5: 14,
            6: 14,
            7: 14,
            8: 22,  # H (mais espaçada)
            9: 12,
            10: 12,
            11: 18
        }
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.column_dimensions["A"].width = 2.5  # margem

        processed += 1
        progress_callback(int(processed / max(1, total_names) * 90), f"Gerando {nome} ({processed}/{total_names})")

    if "tmp" in wb.sheetnames:
        wb.remove(tmp)

    out_path = os.path.join(out_dir, OUT_FILENAME)
    wb.save(out_path)
    progress_callback(100, "Concluído")
    return out_path
